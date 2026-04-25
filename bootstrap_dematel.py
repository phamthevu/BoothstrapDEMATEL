"""
Bootstrap Z-Fuzzy DEMATEL Tool
================================
Đọc file Excel Z-Fuzzy DEMATEL (cấu trúc: sheet DEMATEL + Ans 1..N),
thực hiện Bootstrap để đánh giá độ ổn định, xuất kết quả ra Excel + biểu đồ IRM.
"""

import numpy as np
import pandas as pd
import openpyxl
from openpyxl.styles import (PatternFill, Font, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter
from openpyxl.chart import ScatterChart, Reference, Series
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
from matplotlib.patches import FancyArrowPatch
import re
import sys
import warnings
warnings.filterwarnings('ignore')

# ─── Thang đánh giá Z-Fuzzy (L, M, U) ─────────────────────────────────────
FUZZY_SCALE = {
    "AI;VH": (0.2846, 0.3795, 0.3795),
    "AI;H":  (0.2510, 0.3347, 0.3347),
    "VI;VH": (0.1897, 0.2846, 0.3795),
    "AI;M":  (0.2121, 0.2828, 0.2828),
    "VI;H":  (0.1673, 0.2510, 0.3347),
    "VI;M":  (0.1414, 0.2121, 0.2828),
    "AI;L":  (0.1643, 0.2191, 0.2191),
    "FI;VH": (0.0949, 0.1897, 0.2846),
    "FI;H":  (0.0837, 0.1673, 0.2510),
    "VI;L":  (0.1095, 0.1643, 0.2191),
    "FI;M":  (0.0707, 0.1414, 0.2121),
    "AI;VL": (0.0949, 0.1265, 0.1265),
    "FI;L":  (0.0548, 0.1095, 0.1643),
    "VI;VL": (0.0632, 0.0949, 0.1265),
    "WI;VH": (0.0000, 0.0949, 0.1897),
    "WI;H":  (0.0000, 0.0837, 0.1673),
    "WI;M":  (0.0000, 0.0707, 0.1414),
    "FI;VL": (0.0316, 0.0632, 0.0949),
    "WI;L":  (0.0000, 0.0548, 0.1095),
    "NI;VH": (0.0000, 0.0000, 0.0949),
    "WI;VL": (0.0000, 0.0316, 0.0632),
    "NI;H":  (0.0000, 0.0000, 0.0837),
    "NI;M":  (0.0000, 0.0000, 0.0707),
    "NI;L":  (0.0000, 0.0000, 0.0548),
    "NI;VL": (0.0000, 0.0000, 0.0316),
    "0":     (0.0000, 0.0000, 0.0000),
}

def parse_fuzzy_label(val):
    """Chuyển nhãn fuzzy hoặc số 0 sang tuple (l, m, u)."""
    if val is None or val == 0 or str(val).strip() == "0":
        return (0.0, 0.0, 0.0)
    key = str(val).strip()
    if key in FUZZY_SCALE:
        return FUZZY_SCALE[key]
    return (0.0, 0.0, 0.0)

def defuzzify(l, m, u):
    """Defuzz tam giác:"""
    return (l + m + u) / 3

# ─── Đọc dữ liệu từ các sheet Ans ─────────────────────────────────────────
def read_expert_sheets(filepath):
    """
    Trả về: factors (list), experts (list of 2D np.array defuzzified)
    """
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    sheet_names = wb.sheetnames

    ans_sheets = [s for s in sheet_names if re.match(r'^Ans\s+\d+$', s, re.IGNORECASE)]
    ans_sheets.sort(key=lambda x: int(re.search(r'\d+', x).group()))

    if not ans_sheets:
        print("❌ Không tìm thấy sheet 'Ans N'. Kiểm tra lại tên sheet.")
        sys.exit(1)

    # Lấy tên nhân tố từ sheet Ans 1
    ws0 = wb[ans_sheets[0]]
    rows0 = list(ws0.iter_rows(values_only=True))
    header_row = rows0[0]
    factors = [str(c).strip() for c in header_row[1:] if c is not None]
    n = len(factors)

    experts = []
    for sheet_name in ans_sheets:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        # Lấy 12 hàng dữ liệu (bỏ header row 0)
        matrix = np.zeros((n, n))
        for i in range(n):
            row = rows[i + 1]  # skip header
            for j in range(n):
                val = row[j + 1]  # skip row label
                l, m, u = parse_fuzzy_label(val)
                matrix[i, j] = defuzzify(l, m, u)
        experts.append(matrix)

    wb.close()
    print(f"✅ Đọc xong {len(experts)} chuyên gia, {n} nhân tố: {factors}")
    return factors, experts

# ─── Tính DEMATEL từ một tập expert matrices ────────────────────────────────
def compute_dematel(matrices):
    """
    matrices: list of n×n np.array
    Trả về: r+c (prominence), r-c (relation), r, c, T (total influence matrix)
    """
    n = matrices[0].shape[0]
    # Trung bình ma trận
    A = np.mean(matrices, axis=0)

    # Chuẩn hóa
    row_sums = A.sum(axis=1)
    col_sums = A.sum(axis=0)
    k = max(row_sums.max(), col_sums.max())
    if k == 0:
        return np.zeros(n), np.zeros(n), np.zeros(n), np.zeros(n), A
    D = A / k

    # Ma trận ảnh hưởng tổng thể T = D(I-D)^-1
    I = np.eye(n)
    try:
        T = D @ np.linalg.inv(I - D)
    except np.linalg.LinAlgError:
        T = D @ np.linalg.pinv(I - D)

    r = T.sum(axis=1)  # tổng hàng
    c = T.sum(axis=0)  # tổng cột

    return r + c, r - c, r, c, T

# ─── Bootstrap chính ─────────────────────────────────────────────────────────
def run_bootstrap(factors, experts, B=2000, seed=42, alpha=0.05):
    """
    B vòng bootstrap → CI cho R+C và R-C
    """
    rng = np.random.default_rng(seed)
    n_exp = len(experts)
    n_fac = len(factors)

    prominence_boot = np.zeros((B, n_fac))
    relation_boot   = np.zeros((B, n_fac))

    for b in range(B):
        idx = rng.integers(0, n_exp, size=n_exp)
        sample = [experts[i] for i in idx]
        rc_plus, rc_minus, _, _, _ = compute_dematel(sample)
        prominence_boot[b] = rc_plus
        relation_boot[b]   = rc_minus

    # Kết quả gốc (không bootstrap)
    rc_plus_orig, rc_minus_orig, r_orig, c_orig, T_orig = compute_dematel(experts)

    lo = alpha / 2
    hi = 1 - alpha / 2

    results = []
    for i, fac in enumerate(factors):
        p_mean  = prominence_boot[:, i].mean()
        p_std   = prominence_boot[:, i].std()
        p_lo    = np.percentile(prominence_boot[:, i], lo*100)
        p_hi    = np.percentile(prominence_boot[:, i], hi*100)

        r_mean  = relation_boot[:, i].mean()
        r_std   = relation_boot[:, i].std()
        r_lo    = np.percentile(relation_boot[:, i], lo*100)
        r_hi    = np.percentile(relation_boot[:, i], hi*100)

        # p-value: tỷ lệ bootstrap samples có cùng dấu với gốc
        orig_sign = np.sign(rc_minus_orig[i])
        if orig_sign > 0:
            p_val = (relation_boot[:, i] <= 0).mean()
        elif orig_sign < 0:
            p_val = (relation_boot[:, i] >= 0).mean()
        else:
            p_val = 1.0

        # Ý nghĩa thống kê: CI không chứa 0
        sig = "✓ Có ý nghĩa" if (r_lo > 0 or r_hi < 0) else "✗ Không chắc"
        cause_effect = "Cause" if rc_minus_orig[i] > 0 else "Effect"

        # Trọng số
        w = rc_plus_orig[i] / rc_plus_orig.sum()

        results.append({
            "Nhân tố": fac,
            "r (gốc)": round(r_orig[i], 4),
            "c (gốc)": round(c_orig[i], 4),
            "r+c (gốc)": round(rc_plus_orig[i], 4),
            "r-c (gốc)": round(rc_minus_orig[i], 4),
            "Cause/Effect": cause_effect,
            "Trọng số w": round(w, 4),
            "Xếp hạng": 0,  # sẽ cập nhật sau
            # Bootstrap r+c
            "BS Mean(r+c)": round(p_mean, 4),
            "BS SD(r+c)":   round(p_std, 4),
            f"BS CI_lo(r+c) {int((1-alpha)*100)}%": round(p_lo, 4),
            f"BS CI_hi(r+c) {int((1-alpha)*100)}%": round(p_hi, 4),
            # Bootstrap r-c
            "BS Mean(r-c)": round(r_mean, 4),
            "BS SD(r-c)":   round(r_std, 4),
            f"BS CI_lo(r-c) {int((1-alpha)*100)}%": round(r_lo, 4),
            f"BS CI_hi(r-c) {int((1-alpha)*100)}%": round(r_hi, 4),
            # Kiểm định
            "p-value": round(p_val, 4),
            "Có ý nghĩa TK?": sig,
        })

    # Xếp hạng theo r+c giảm dần
    df = pd.DataFrame(results)
    df["Xếp hạng"] = df["r+c (gốc)"].rank(ascending=False).astype(int)
    df = df.sort_values("Xếp hạng")

    return df, rc_plus_orig, rc_minus_orig, T_orig, prominence_boot, relation_boot

# ─── Xuất Excel đẹp ──────────────────────────────────────────────────────────
def style_header(ws, row, cols, fill_hex, font_color="FFFFFF", font_size=11):
    fill = PatternFill("solid", fgColor=fill_hex)
    font = Font(bold=True, color=font_color, size=font_size, name="Arial")
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

def style_data_row(ws, row, n_cols, is_alt):
    fill = PatternFill("solid", fgColor="F0F4FF" if is_alt else "FFFFFF")
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.font = Font(name="Arial", size=10)

def add_thin_border(ws, min_row, max_row, min_col, max_col):
    thin = Side(style='thin', color="CCCCCC")
    for row in ws.iter_rows(min_row=min_row, max_row=max_row,
                             min_col=min_col, max_col=max_col):
        for cell in row:
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

def export_excel(df, factors, T_orig, prominence_boot, relation_boot,
                 output_path, B, alpha):
    wb = openpyxl.Workbook()

    # ── Sheet 1: Kết quả Bootstrap ─────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Bootstrap DEMATEL"
    ws1.sheet_view.showGridLines = False

    # Tiêu đề lớn
    ws1.merge_cells("A1:V1")
    title_cell = ws1["A1"]
    title_cell.value = f"KẾT QUẢ BOOTSTRAP Z-FUZZY DEMATEL  (B={B:,} lần, CI={(1-alpha)*100:.0f}%)"
    title_cell.font = Font(bold=True, size=14, color="FFFFFF", name="Arial")
    title_cell.fill = PatternFill("solid", fgColor="1A3A6B")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[1].height = 32

    # Header
    headers = list(df.columns)
    n_cols = len(headers)
    for col_idx, h in enumerate(headers, 1):
        ws1.cell(row=2, column=col_idx, value=h)
    style_header(ws1, 2, n_cols, "2E5FA3")
    ws1.row_dimensions[2].height = 28

    # Màu theo Cause/Effect
    cause_fill = PatternFill("solid", fgColor="DFF0D8")
    effect_fill = PatternFill("solid", fgColor="F2DEDE")
    sig_fill = PatternFill("solid", fgColor="FFF3CD")

    for row_idx, (_, row_data) in enumerate(df.iterrows(), 3):
        is_alt = (row_idx % 2 == 0)
        style_data_row(ws1, row_idx, n_cols, is_alt)
        for col_idx, val in enumerate(row_data, 1):
            cell = ws1.cell(row=row_idx, column=col_idx, value=val)
            # Màu đặc biệt cho Cause/Effect
            if headers[col_idx - 1] == "Cause/Effect":
                cell.fill = cause_fill if val == "Cause" else effect_fill
                cell.font = Font(bold=True, name="Arial", size=10,
                                 color="2D6A2D" if val == "Cause" else "A94442")
            if headers[col_idx - 1] == "Có ý nghĩa TK?":
                if "Không" in str(val):
                    cell.fill = sig_fill
                    cell.font = Font(bold=True, name="Arial", size=10, color="8A6D3B")

    add_thin_border(ws1, 2, 2 + len(df), 1, n_cols)

    # Độ rộng cột
    col_widths = [12, 10, 10, 12, 12, 13, 12, 10,
                  14, 12, 18, 18, 14, 12, 18, 18,
                  10, 18]
    for i, w in enumerate(col_widths[:n_cols], 1):
        ws1.column_dimensions[get_column_letter(i)].width = w

    # Ghi chú
    note_row = 3 + len(df) + 1
    ws1.cell(row=note_row, column=1,
             value=f"📌 Ghi chú: B={B:,} bootstrap samples | CI={int((1-alpha)*100)}% Percentile Bootstrap | "
                   f"p-value = tỷ lệ mẫu bootstrap trái dấu gốc").font = Font(
        italic=True, size=9, color="555555", name="Arial")

    # ── Sheet 2: Ma trận T (TIM) ───────────────────────────────────────────
    ws2 = wb.create_sheet("Ma trận TIM")
    ws2.sheet_view.showGridLines = False

    ws2.merge_cells(f"A1:{get_column_letter(len(factors)+1)}1")
    t2 = ws2["A1"]
    t2.value = "MA TRẬN ẢNH HƯỞNG TỔNG THỂ (Total Influence Matrix - T)"
    t2.font = Font(bold=True, size=13, color="FFFFFF", name="Arial")
    t2.fill = PatternFill("solid", fgColor="1A3A6B")
    t2.alignment = Alignment(horizontal="center", vertical="center")

    ws2.cell(row=2, column=1, value="")
    for j, fac in enumerate(factors, 2):
        ws2.cell(row=2, column=j, value=fac)
    style_header(ws2, 2, len(factors)+1, "2E5FA3")

    max_t = T_orig.max()
    for i, fac in enumerate(factors):
        ws2.cell(row=i+3, column=1, value=fac).font = Font(bold=True, name="Arial", size=10)
        ws2.cell(row=i+3, column=1).fill = PatternFill("solid", fgColor="2E5FA3")
        ws2.cell(row=i+3, column=1).font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
        for j in range(len(factors)):
            val = T_orig[i, j]
            cell = ws2.cell(row=i+3, column=j+2, value=round(val, 4))
            # Heatmap màu xanh
            intensity = int(255 - (val / max_t) * 180) if max_t > 0 else 255
            hex_col = f"{'%02X' % intensity}{'%02X' % intensity}FF"
            cell.fill = PatternFill("solid", fgColor=hex_col)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.font = Font(name="Arial", size=9)

    add_thin_border(ws2, 2, 2+len(factors), 1, len(factors)+1)
    for col in range(1, len(factors)+2):
        ws2.column_dimensions[get_column_letter(col)].width = 10

    # ── Sheet 3: Bootstrap Distribution ───────────────────────────────────
    ws3 = wb.create_sheet("BS Distribution")
    ws3.sheet_view.showGridLines = False

    ws3.merge_cells("A1:F1")
    d3 = ws3["A1"]
    d3.value = "THỐNG KÊ PHÂN PHỐI BOOTSTRAP (r-c)"
    d3.font = Font(bold=True, size=13, color="FFFFFF", name="Arial")
    d3.fill = PatternFill("solid", fgColor="1A3A6B")
    d3.alignment = Alignment(horizontal="center", vertical="center")

    bs_headers = ["Nhân tố", "Min", "P10", "P25", "Median", "P75", "P90", "Max", "Mean", "SD"]
    for ci, h in enumerate(bs_headers, 1):
        ws3.cell(row=2, column=ci, value=h)
    style_header(ws3, 2, len(bs_headers), "2E5FA3")

    for i, fac in enumerate(factors):
        rd = relation_boot[:, i]
        row_data = [fac,
                    round(rd.min(), 4), round(np.percentile(rd, 10), 4),
                    round(np.percentile(rd, 25), 4), round(np.median(rd), 4),
                    round(np.percentile(rd, 75), 4), round(np.percentile(rd, 90), 4),
                    round(rd.max(), 4), round(rd.mean(), 4), round(rd.std(), 4)]
        for ci, val in enumerate(row_data, 1):
            cell = ws3.cell(row=i+3, column=ci, value=val)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.font = Font(name="Arial", size=10)
            if ci > 1:
                if isinstance(val, float):
                    color = "DFF0D8" if val > 0 else "F2DEDE" if val < 0 else "FFFFFF"
                    cell.fill = PatternFill("solid", fgColor=color)

    add_thin_border(ws3, 2, 2+len(factors), 1, len(bs_headers))
    for ci, w in enumerate([14,8,8,8,10,8,8,8,10,8], 1):
        ws3.column_dimensions[get_column_letter(ci)].width = w

    wb.save(output_path)
    print(f"✅ Đã xuất Excel: {output_path}")

# ─── Vẽ biểu đồ IRM với Bootstrap CI ─────────────────────────────────────
def plot_irm(df, factors, prominence_boot, relation_boot, output_path, alpha):
    fig, axes = plt.subplots(1, 2, figsize=(18, 8))
    fig.patch.set_facecolor('#F8F9FA')

    # ── Biểu đồ 1: IRM gốc + CI ellipse ───────────────────────────────────
    ax = axes[0]
    ax.set_facecolor('#FAFAFA')

    x = df["r+c (gốc)"].values
    y = df["r-c (gốc)"].values
    ci_col_lo = f"BS CI_lo(r-c) {int((1-alpha)*100)}%"
    ci_col_hi = f"BS CI_hi(r-c) {int((1-alpha)*100)}%"
    y_lo = df[ci_col_lo].values
    y_hi = df[ci_col_hi].values

    cause_color  = "#2196F3"
    effect_color = "#F44336"
    sig_marker   = "o"

    for i, fac in enumerate(df["Nhân tố"].values):
        is_cause = df["Cause/Effect"].values[i] == "Cause"
        is_sig   = "✓" in str(df["Có ý nghĩa TK?"].values[i])
        color = cause_color if is_cause else effect_color
        marker = sig_marker

        # Vẽ CI bar
        ax.plot([x[i], x[i]], [y_lo[i], y_hi[i]],
                color=color, alpha=0.3, linewidth=2, zorder=1)
        ax.plot([x[i]-0.005, x[i]+0.005], [y_lo[i], y_lo[i]],
                color=color, alpha=0.4, linewidth=1.5, zorder=1)
        ax.plot([x[i]-0.005, x[i]+0.005], [y_hi[i], y_hi[i]],
                color=color, alpha=0.4, linewidth=1.5, zorder=1)

        # Điểm
        ax.scatter(x[i], y[i], c=color, s=120 if is_sig else 70,
                   zorder=3, edgecolors='white', linewidth=1.5,
                   alpha=0.9, marker=marker if is_sig else "D")

        # Nhãn
        ax.annotate(fac, (x[i], y[i]),
                    textcoords="offset points", xytext=(6, 4),
                    fontsize=9, fontweight='bold', color=color,
                    bbox=dict(boxstyle="round,pad=0.2", fc="white", alpha=0.7, ec=color, lw=0.8))

    # Đường phân chia Cause/Effect
    ax.axhline(0, color='#555555', linewidth=1.2, linestyle='--', alpha=0.7, zorder=2)

    # Đường prominence trung bình
    x_thresh = x.mean()
    ax.axvline(x_thresh, color='#9E9E9E', linewidth=1, linestyle=':', alpha=0.6, zorder=2)

    # Vùng
    y_lim = ax.get_ylim()
    x_lim = ax.get_xlim()
    ax.fill_between([x_lim[0], x_thresh], 0, y_lim[1], alpha=0.04, color='blue')
    ax.fill_between([x_thresh, x_lim[1]], 0, y_lim[1], alpha=0.06, color='blue')
    ax.fill_between([x_lim[0], x_thresh], y_lim[0], 0, alpha=0.04, color='red')
    ax.fill_between([x_thresh, x_lim[1]], y_lim[0], 0, alpha=0.06, color='red')

    ax.text(x_lim[1]*0.98, y_lim[1]*0.92, "CAUSE", ha='right', va='top',
            fontsize=11, color=cause_color, fontweight='bold', alpha=0.5)
    ax.text(x_lim[1]*0.98, y_lim[0]*0.92, "EFFECT", ha='right', va='bottom',
            fontsize=11, color=effect_color, fontweight='bold', alpha=0.5)

    ax.set_xlabel("r + c  (Prominence / Mức độ ảnh hưởng)", fontsize=11, labelpad=8)
    ax.set_ylabel("r − c  (Relation / Nguyên nhân - Hệ quả)", fontsize=11, labelpad=8)
    ax.set_title("Impact-Relation Map (IRM)\nvới Bootstrap 95% CI cho r−c",
                 fontsize=12, fontweight='bold', pad=12)
    ax.grid(True, alpha=0.3, linestyle='--')

    legend_elements = [
        mpatches.Patch(color=cause_color, label='Cause (Nguyên nhân)'),
        mpatches.Patch(color=effect_color, label='Effect (Hệ quả)'),
        plt.scatter([], [], marker='o', c='gray', s=80, label='Có ý nghĩa TK'),
        plt.scatter([], [], marker='D', c='gray', s=60, label='Chưa rõ ràng'),
    ]
    ax.legend(handles=legend_elements, loc='upper left', fontsize=9,
              framealpha=0.8, edgecolor='#CCCCCC')

    # ── Biểu đồ 2: Bootstrap violin plot cho r-c ───────────────────────────
    ax2 = axes[1]
    ax2.set_facecolor('#FAFAFA')

    fac_labels = df["Nhân tố"].values
    n_fac = len(fac_labels)
    # Sắp xếp theo r-c gốc giảm dần
    order = np.argsort(df["r-c (gốc)"].values)[::-1]
    boot_data = [relation_boot[:, df["Nhân tố"].values.tolist().index(fac_labels[i])]
                 for i in order]
    labels_ord = [fac_labels[i] for i in order]
    orig_ord   = [df["r-c (gốc)"].values[i] for i in order]
    cause_ord  = [df["Cause/Effect"].values[i] for i in order]

    colors = [cause_color if c == "Cause" else effect_color for c in cause_ord]
    vp = ax2.violinplot(boot_data, positions=range(n_fac),
                        showmedians=True, showextrema=False)

    for i, (body, color) in enumerate(zip(vp['bodies'], colors)):
        body.set_facecolor(color)
        body.set_alpha(0.35)
        body.set_edgecolor(color)

    vp['cmedians'].set_color('#333333')
    vp['cmedians'].set_linewidth(2)

    # Vẽ điểm gốc
    ax2.scatter(range(n_fac), orig_ord, c=colors, s=60, zorder=5,
                edgecolors='white', linewidth=1.5)

    ax2.axhline(0, color='#555555', linewidth=1.2, linestyle='--', alpha=0.7)
    ax2.set_xticks(range(n_fac))
    ax2.set_xticklabels(labels_ord, rotation=45, ha='right', fontsize=9)
    ax2.set_ylabel("r − c (Bootstrap Distribution)", fontsize=11)
    ax2.set_title("Phân phối Bootstrap của r−c\n(Violin Plot - sắp xếp theo r−c gốc)",
                  fontsize=12, fontweight='bold', pad=12)
    ax2.grid(True, alpha=0.3, linestyle='--', axis='y')

    fig.suptitle("Bootstrap Z-Fuzzy DEMATEL Analysis", fontsize=14,
                 fontweight='bold', y=1.01, color='#1A3A6B')
    plt.tight_layout()
    plt.savefig(output_path, dpi=180, bbox_inches='tight',
                facecolor=fig.get_facecolor())
    plt.close()
    print(f"✅ Đã xuất biểu đồ IRM: {output_path}")

# ─── Main ────────────────────────────────────────────────────────────────────
def main():
    INPUT_FILE  = "INPUT.xlsx"
    OUTPUT_XLS  = "OUTPUT.xlsx"
    OUTPUT_IMG  = "OUTPUT.png"
    B     = 5000
    ALPHA = 0.05
    SEED  = 80

    print("=" * 60)
    print("  Bootstrap Z-Fuzzy DEMATEL Tool")
    print("=" * 60)

    # 1. Đọc dữ liệu
    factors, experts = read_expert_sheets(INPUT_FILE)

    # 2. Bootstrap
    print(f"\n🔄 Đang chạy Bootstrap (B={B:,} lần)...")
    df, rc_plus, rc_minus, T_orig, p_boot, r_boot = run_bootstrap(
        factors, experts, B=B, seed=SEED, alpha=ALPHA)

    # 3. Xuất Excel
    print("\n📊 Xuất kết quả Excel...")
    export_excel(df, factors, T_orig, p_boot, r_boot, OUTPUT_XLS, B, ALPHA)

    # 4. Vẽ IRM
    print("🎨 Vẽ biểu đồ IRM...")
    plot_irm(df, factors, p_boot, r_boot, OUTPUT_IMG, ALPHA)

    # 5. In tóm tắt
    print("\n" + "=" * 60)
    print("  TÓM TẮT KẾT QUẢ")
    print("=" * 60)
    display_cols = ["Nhân tố", "r+c (gốc)", "r-c (gốc)", "Cause/Effect",
                    f"BS CI_lo(r-c) {int((1-ALPHA)*100)}%",
                    f"BS CI_hi(r-c) {int((1-ALPHA)*100)}%",
                    "p-value", "Có ý nghĩa TK?"]
    print(df[display_cols].to_string(index=False))
    print(f"\n  Số chuyên gia: {len(experts)}")
    print(f"  Số nhân tố:   {len(factors)}")
    print(f"  B (Bootstrap): {B:,}")
    print(f"  Confidence:    {int((1-ALPHA)*100)}%")
    print("=" * 60)

if __name__ == "__main__":
    main()

def run_pipeline(input_file, output_xls, output_img, B=5000, alpha=0.05, seed=80):
    print("🚀 Running pipeline...")

    factors, experts = read_expert_sheets(input_file)

    df, rc_plus, rc_minus, T_orig, p_boot, r_boot = run_bootstrap(
        factors, experts, B=B, seed=seed, alpha=alpha)

    export_excel(df, factors, T_orig, p_boot, r_boot, output_xls, B, alpha)

    plot_irm(df, factors, p_boot, r_boot, output_img, alpha)

    return df

if __name__ == "__main__":
    main()
